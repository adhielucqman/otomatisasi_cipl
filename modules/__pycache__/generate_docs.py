# import pandas as pd
# import os
# from openpyxl import load_workbook
# from tqdm import tqdm
# from copy import copy
# from num2words import num2words

# def generate_ci_pl():
#     print("[INFO] Membaca file Master Order List dan Guideline Shipment...")

#     # --- [FITUR BARU] KAMUS HS CODE ---
#     # Tambahkan Style baru beserta HS Code-nya di sini
#     hs_kamus = {
#         "211354": "640419",
#         # "StyleLain": "123456",  <-- contoh cara menambahkan di masa depan
#     }
#     # ----------------------------------

#     try:
#         # 1. Membaca Master Order List & Guideline Shipment
#         df = pd.read_excel("Master Order List Adi.xlsx")
#         df_db = pd.read_excel("guideline_shipment.xlsx")
#     except Exception as e:
#         print(f"[ERROR] Gagal membaca file excel: {e}")
#         return

#     # 2. Membuat folder utama Output jika belum ada
#     output_dir = "Output_FI_PL_SI"
#     os.makedirs(output_dir, exist_ok=True)

#     # 3. GRUP PERTAMA: Mengelompokkan berdasarkan TL# (Untuk membuat Folder Gabungan)
#     grouped_tl = df.groupby("TL#")
    
#     print(f"[INFO] Menemukan {len(grouped_tl)} grup TL#. Memulai proses generate...\n")

#     # 4. Looping Utama: Proses per Grup TL
#     for tl_num, tl_group in tqdm(grouped_tl, desc="Proses Group TL"):
        
#         # -- TENTUKAN NAMA FOLDER GABUNGAN --
#         # Mengambil CI# unik, diurutkan agar rapi (cth: 84 & 85), lalu digabung dengan " & "
#         ci_list = sorted(tl_group["CI#"].unique().astype(str))
#         all_ci_names = " & ".join(ci_list)
        
#         # Mengambil data Code TL untuk penamaan folder (cth: CFS-AU-025)
#         tl_code_folder = tl_group.iloc[0]["TL#"]
        
#         # Format Folder: [CI1 & CI2] - [Code TL]
#         folder_name = f"{all_ci_names} - {tl_code_folder}".replace("/", "-") 
#         folder_path = os.path.join(output_dir, folder_name)
#         os.makedirs(folder_path, exist_ok=True)
        
#         # 5. SUB-LOOPING: Memproses masing-masing Invoice di dalam grup TL yang sama
#         grouped_inv = tl_group.groupby("Factory Inv#")
        
#         for inv_num, group in grouped_inv:
#             # -- VARIABLE DATA HEADER INVOICE, PACKING LIST & FiNAL SI --
#             first_row = group.iloc[0]
#             date_export = first_row["EXII-FTY"]
#             country = first_row["Customer Country"]
#             kode_negara = first_row["Code Country"]
#             tl_code = first_row["Code TL"]
#             te_inv = first_row["CI#"]
#             ship_by = first_row["Ship By"]
#             ship_mode = first_row["MOT"]
#             branch_plant = first_row["Branch Plant"]
#             pod = first_row["POD"]
#             no_pen_peb = first_row["NO PEN PEB"]
#             peb_date = first_row["PEB DATE"]
#             type_cont_truck = first_row["Ctnr type/Truck type"]
#             carrier = first_row["FWDR"]
#             contNo_sealNo = first_row["Detail Cntainer no./Seal/truck no"]
#             booking_no = first_row["Booking No / SO No."]


#             # --- MENCARI CONSIGNEE & DELIVERY ADDRESS DI DATABASE ---
#             match = df_db[
#                 (df_db["Country"] == kode_negara) & 
#                 (df_db["By"] == ship_by) & 
#                 (df_db["Branch Plant"] == branch_plant) &
#                 (df_db["POD"] == pod)
#             ]
            
#             if not match.empty:
#                 consignee_data = match.iloc[0]["Consignee"]
#                 delivery_data = match.iloc[0]["Notify Party/Delivery Address"]
#             else:
#                 consignee_data = "DATA TIDAK DITEMUKAN"
#                 delivery_data = "DATA TIDAK DITEMUKAN"
#                 print(f"\n[WARNING] Consignee tidak ditemukan untuk Inv {inv_num}")

#             # ==========================================
#             # PROSES 1: COMMERCIAL INVOICE (CI)
#             # ==========================================
#             wb_ci = load_workbook("Template_CI_V2.xlsx")
#             ws_ci = wb_ci.active
            
#             # Isi Header CI
#             ws_ci["E6"] = inv_num
#             ws_ci["G6"] = date_export
#             ws_ci["G10"] = country
#             ws_ci["A19"] = ship_mode
#             ws_ci["E19"] = ship_by
#             ws_ci["E21"] = pod
            
#             # Isi Consignee (Dipecah per baris mulai dari C12)
#             if isinstance(consignee_data, str):
#                 for i, baris in enumerate(consignee_data.split('\n')):
#                     ws_ci[f"C{13 + i}"] = baris.strip()
#             else:
#                 ws_ci["C13"] = consignee_data

#             # Isi Delivery Address (Dipecah per baris mulai dari E12)
#             if isinstance(delivery_data, str):
#                 for i, baris in enumerate(delivery_data.split('\n')):
#                     ws_ci[f"E{13 + i}"] = baris.strip()
#             else:
#                 ws_ci["E13"] = delivery_data
            
#             # Isi Detail Barang CI & PL
#             start_row = 26
#             num_items = len(group)
            
#             # Insert Row & Copy Style jika barang lebih dari 1
#             if num_items > 1:
#                 ws_ci.insert_rows(start_row + 1, amount=num_items - 1)
#                 for r in range(start_row + 1, start_row + num_items):
#                     for c in range(1, 10): 
#                         source_cell = ws_ci.cell(row=start_row, column=c)
#                         target_cell = ws_ci.cell(row=r, column=c)
#                         if source_cell.has_style:
#                             target_cell.font = copy(source_cell.font)
#                             target_cell.border = copy(source_cell.border)
#                             target_cell.fill = copy(source_cell.fill)
#                             target_cell.number_format = copy(source_cell.number_format)
#                             target_cell.alignment = copy(source_cell.alignment)

#             total_ctn, total_pairs, total_amount = 0, 0, 0
            
#             for idx, (_, row) in enumerate(group.iterrows()):
#                 curr_row = start_row + idx
#                 ws_ci[f"A{curr_row}"] = row["PO#"]
#                 ws_ci[f"B{curr_row}"] = row["ProductNO"]
#                 ws_ci[f"C{curr_row}"] = row["EnglishName"]
#                 # --- [FITUR BARU] CEK HS CODE DARI KAMUS ---
#                 # Mengambil teks aslinya (Misal: "211354-001")
#                 full_product_no = str(row["ProductNO"]).strip()
                
#                 # Memotong teks berdasarkan tanda "-" dan mengambil bagian depannya saja (Index 0)
#                 # "211354-001" -> menjadi "211354"
#                 style_only = full_product_no.split('-')[0].strip()
                
#                 # Mencari "211354" di dalam kamus
#                 ws_ci[f"D{curr_row}"] = hs_kamus.get(style_only, "640299") 
#                 # -------------------------------------------
#                 ws_ci[f"E{curr_row}"] = row["CTN"]
#                 ws_ci[f"F{curr_row}"] = row["PAIRS"]
#                 ws_ci[f"G{curr_row}"] = row["TE Price"]
#                 ws_ci[f"H{curr_row}"] = row["SCI Amount"]
                
#                 total_ctn += row["CTN"]
#                 total_pairs += row["PAIRS"]
#                 total_amount += row["SCI Amount"]
                
#             # Tulis Total (Bawah Tabel)
#             last_row_ci = start_row + num_items + 1
#             ws_ci[f"E{last_row_ci}"] = total_ctn
#             ws_ci[f"F{last_row_ci}"] = total_pairs
#             ws_ci[f"H{last_row_ci}"] = total_amount

#             # --- [FITUR TERBILANG & MENGISI TOTAL CI DI SAMPING] ---
#             dollars = int(total_amount)
#             cents = int(round((total_amount - dollars) * 100))
#             terbilang_dollars = num2words(dollars, lang='en').replace('-', ' ').replace(',', '').upper()
            
#             if cents > 0:
#                 terbilang_cents = num2words(cents, lang='en').replace('-', ' ').upper()
#                 terbilang_text = f"Total Amount Say US Dollar {terbilang_dollars} AND CENTS {terbilang_cents} ONLY"
#             else:
#                 terbilang_text = f"Total Amount Say US Dollar {terbilang_dollars} ONLY"
                
#             for r in range(start_row, ws_ci.max_row + 1):
#                 cell_val = ws_ci[f"A{r}"].value
#                 if isinstance(cell_val, str):
#                     teks = cell_val.strip().lower()
#                     if teks == "total cartons":
#                         ws_ci[f"B{r}"] = total_ctn
#                     elif teks == "total quantity":
#                         ws_ci[f"B{r}"] = total_pairs
#                     elif teks == "total amount":
#                         ws_ci[f"B{r}"] = total_amount
#                     elif "total amount say" in teks or "say us dollar" in teks:
#                         ws_ci[f"A{r}"] = terbilang_text
#             # --------------------------------------------------------

#             # Atur Page Break & Print Area CI
#             max_row_ci = ws_ci.max_row
#             ws_ci.print_area = f"A1:H{max_row_ci - 5}"
#             ws_ci.sheet_properties.pageSetUpPr.fitToPage = True
#             ws_ci.page_setup.fitToWidth = 1
#             ws_ci.page_setup.fitToHeight = 0
            
#             # Simpan file CI di dalam Folder Gabungan
#             file_name_ci = f"{inv_num} INV {tl_code}.xlsx".replace("/", "-")
#             wb_ci.save(os.path.join(folder_path, file_name_ci))
            
            
#             # ==========================================
#             # PROSES 2: PACKING LIST (PL)
#             # ==========================================
#             wb_pl = load_workbook("Template_PL_V2.xlsx")
#             ws_pl = wb_pl.active
            
#             # Isi Header PL
#             ws_pl["E6"] = inv_num
#             ws_pl["G6"] = date_export
#             ws_pl["G10"] = country
#             ws_pl["A19"] = ship_mode
#             ws_pl["E19"] = ship_by
#             ws_pl["E21"] = pod
            
#             # Isi Consignee (Dipecah per baris mulai dari C12)
#             if isinstance(consignee_data, str):
#                 for i, baris in enumerate(consignee_data.split('\n')):
#                     ws_pl[f"C{13 + i}"] = baris.strip()
#             else:
#                 ws_pl["C13"] = consignee_data

#             # Isi Delivery Address (Dipecah per baris mulai dari E12)
#             if isinstance(delivery_data, str):
#                 for i, baris in enumerate(delivery_data.split('\n')):
#                     ws_pl[f"E{13 + i}"] = baris.strip()
#             else:
#                 ws_pl["E13"] = delivery_data

#             # Insert Row & Copy Style jika barang lebih dari 1
#             if num_items > 1:
#                 ws_pl.insert_rows(start_row + 1, amount=num_items - 1)
#                 for r in range(start_row + 1, start_row + num_items):
#                     for c in range(1, 10): 
#                         source_cell = ws_pl.cell(row=start_row, column=c)
#                         target_cell = ws_pl.cell(row=r, column=c)
#                         if source_cell.has_style:
#                             target_cell.font = copy(source_cell.font)
#                             target_cell.border = copy(source_cell.border)
#                             target_cell.fill = copy(source_cell.fill)
#                             target_cell.number_format = copy(source_cell.number_format)
#                             target_cell.alignment = copy(source_cell.alignment)
            
#             # Isi Detail Barang PL
#             total_nw, total_gw, total_cbm = 0, 0, 0
            
#             for idx, (_, row) in enumerate(group.iterrows()):
#                 curr_row = start_row + idx
#                 ws_pl[f"A{curr_row}"] = row["PO#"]
#                 ws_pl[f"B{curr_row}"] = row["ProductNO"]
#                 ws_pl[f"C{curr_row}"] = row["EnglishName"]
#                 # --- [FITUR BARU] CEK HS CODE DARI KAMUS ---
#                 full_product_no = str(row["ProductNO"]).strip()
#                 style_only = full_product_no.split('-')[0].strip()
#                 ws_pl[f"D{curr_row}"] = hs_kamus.get(style_only, "640299")
#                 # -------------------------------------------
#                 ws_pl[f"E{curr_row}"] = row["CTN"]
#                 ws_pl[f"F{curr_row}"] = row["PAIRS"]
#                 ws_pl[f"G{curr_row}"] = row["N.W"]
#                 ws_pl[f"H{curr_row}"] = row["G.W"]
#                 ws_pl[f"I{curr_row}"] = row["CBM"]
                
#                 total_nw += row["N.W"]
#                 total_gw += row["G.W"]
#                 total_cbm += row["CBM"]
                
#             # Tulis Total (Bawah Tabel)
#             last_row_pl = start_row + num_items + 1
#             ws_pl[f"E{last_row_pl}"] = total_ctn      
#             ws_pl[f"F{last_row_pl}"] = total_pairs    
#             ws_pl[f"G{last_row_pl}"] = total_nw
#             ws_pl[f"H{last_row_pl}"] = total_gw
#             ws_pl[f"I{last_row_pl}"] = total_cbm

#             # --- [FITUR MENGISI TOTAL PL DI SAMPING] ---
#             for r in range(start_row, ws_pl.max_row + 1):
#                 cell_val = ws_pl[f"A{r}"].value
#                 if isinstance(cell_val, str):
#                     teks = cell_val.strip().lower()
#                     if teks == "total cartons":
#                         ws_pl[f"B{r}"] = total_ctn
#                     elif teks == "total quantity":
#                         ws_pl[f"B{r}"] = total_pairs
#                     elif teks == "total net weight":
#                         ws_pl[f"B{r}"] = total_nw
#                     elif teks == "total gross weight":
#                         ws_pl[f"B{r}"] = total_gw
#                     elif teks == "total cbm":
#                         ws_pl[f"B{r}"] = total_cbm
#             # -------------------------------------------

#             # Atur Page Break & Print Area PL
#             max_row_pl = ws_pl.max_row
#             ws_pl.print_area = f"A1:I{max_row_pl - 4}"
#             ws_pl.sheet_properties.pageSetUpPr.fitToPage = True
#             ws_pl.page_setup.fitToWidth = 1
#             ws_pl.page_setup.fitToHeight = 0
            
#             # Simpan file PL di dalam Folder Gabungan
#             file_name_pl = f"{inv_num} PL {tl_code}.xlsx".replace("/", "-")
#             wb_pl.save(os.path.join(folder_path, file_name_pl))

#         print("\n[INFO] CI & PL Selesai. Memulai proses generate FINAL SI (FSI)...")

#     # ==========================================
#     # PROSES 3: FINAL SI (FSI) - GLOBAL BERDASARKAN BOOKING NO
#     # ==========================================
#     # Mengelompokkan data induk secara langsung berdasarkan Booking No
#     grouped_booking_global = df.groupby("Booking No / SO No.", sort=False)
    
#     for book_no, book_group in tqdm(grouped_booking_global, desc="Proses FSI per Booking"):
#         wb_si = load_workbook("Template_FSI_V2.xlsx")
        
#         if "CONTAINER" in wb_si.sheetnames:
#             ws_base = wb_si["CONTAINER"]
#         else:
#             ws_base = wb_si.active 

#         start_row_si = 9
        
#         # Di dalam 1 Booking No (yang mungkin terdiri dari banyak TL#), kelompokkan per Invoice
#         grouped_inv_fsi = book_group.groupby("Factory Inv#", sort=False)
        
#         # Siapkan sheet CONTAINER sebanyak jumlah invoice unik
#         sheets_for_inv = []
#         for inv_idx in range(len(grouped_inv_fsi)):
#             if inv_idx == 0:
#                 sheets_for_inv.append(ws_base)
#             else:
#                 new_ws = wb_si.copy_worksheet(ws_base)
#                 new_ws.title = f"CONTAINER {inv_idx + 1}"
#                 sheets_for_inv.append(new_ws)
        
#         # Looping mengisi masing-masing sheet
#         for inv_idx, (inv_num, inv_group) in enumerate(grouped_inv_fsi):
#             ws_current = sheets_for_inv[inv_idx]
#             num_items_inv = len(inv_group)
            
#             # Ambil variabel header khusus untuk invoice ini
#             first_row_inv = inv_group.iloc[0]
#             date_export = first_row_inv["EXII-FTY"]
#             carrier = first_row_inv["FWDR"]
#             contNo_sealNo = first_row_inv["Detail Cntainer no./Seal/truck no"]
#             type_cont_truck = first_row_inv["Ctnr type/Truck type"]
#             te_inv = first_row_inv["CI#"]
#             no_pen_peb = first_row_inv["NO PEN PEB"]
#             peb_date = first_row_inv["PEB DATE"]
#             pod = first_row_inv["POD"]
            
#             # 1. Insert Row & Copy Style (Range Kolom C sampai J)
#             if num_items_inv > 1:
#                 ws_current.insert_rows(start_row_si + 1, amount=num_items_inv - 1)
#                 for r in range(start_row_si + 1, start_row_si + num_items_inv):
#                     for c in range(3, 11): 
#                         source_cell = ws_current.cell(row=start_row_si, column=c)
#                         target_cell = ws_current.cell(row=r, column=c)
#                         if source_cell.has_style:
#                             target_cell.font = copy(source_cell.font)
#                             target_cell.border = copy(source_cell.border)
#                             target_cell.fill = copy(source_cell.fill)
#                             target_cell.number_format = copy(source_cell.number_format)
#                             target_cell.alignment = copy(source_cell.alignment)
            
#             # 2. Isi Detail Barang
#             total_ctn_si, total_pairs_si, total_gw_si, total_cbm_si = 0, 0, 0, 0
            
#             for idx, (_, row) in enumerate(inv_group.iterrows()):
#                 curr_row = start_row_si + idx
#                 ws_current[f"C{curr_row}"] = row["PO#"]
#                 ws_current[f"D{curr_row}"] = row["ProductNO"]
#                 ws_current[f"E{curr_row}"] = row["CTN"]
#                 ws_current[f"F{curr_row}"] = row["PAIRS"]
#                 ws_current[f"G{curr_row}"] = row["G.W"]
#                 ws_current[f"H{curr_row}"] = row["CBM"]
#                 ws_current[f"I{curr_row}"] = row["Branch Plant"]
                
#                 full_product_no = str(row["ProductNO"]).strip()
#                 style_only = full_product_no.split('-')[0].strip()
#                 ws_current[f"J{curr_row}"] = hs_kamus.get(style_only, "640299") 
                
#                 total_ctn_si += row["CTN"]
#                 total_pairs_si += row["PAIRS"]
#                 total_gw_si += row["G.W"]
#                 total_cbm_si += row["CBM"]
            
#             # 3. Tulis Total di Bawah Tabel
#             last_row_si = start_row_si + num_items_inv + 6
#             ws_current[f"E{last_row_si}"] = total_ctn_si     
#             ws_current[f"F{last_row_si}"] = total_pairs_si   
#             ws_current[f"G{last_row_si}"] = total_gw_si
#             ws_current[f"H{last_row_si}"] = total_cbm_si
            
#             # 4. Isi Header FSI
#             ws_current["A1"] = f"CARRIER BOOKING# : {book_no}"
#             ws_current["A4"] = inv_num
#             ws_current["A6"] = carrier
#             ws_current["A8"] = contNo_sealNo
#             ws_current["B8"] = type_cont_truck
#             ws_current["A9"] = te_inv
#             ws_current["A10"] = f"EMPTY PICK-UP DATE : {pd.to_datetime(date_export).strftime('%d-%b-%y')}"
#             ws_current["A11"] = f"STUFFING DATE : {pd.to_datetime(date_export).strftime('%d-%b-%y')}"
#             ws_current["A12"] = f"PEB : {no_pen_peb} / {pd.to_datetime(peb_date).strftime('%d-%b-%y')}"
#             ws_current["A13"] = f"POD : {pod}"
#             ws_current["A14"] = f"Final Destination : {pod}"
        
#         # Simpan file Final SI di folder utama (output_dir)
#         file_name_si = f"FINAL SI BC #{book_no}.xlsx".replace("/", "-")
#         wb_si.save(os.path.join(output_dir, file_name_si))

#     print("\n[SUCCESS] Semua data CI, PL & FSI berhasil digenerate dan dikelompokkan dalam subfolder!")

# # Menjalankan fungsi
# if __name__ == "__main__":
#     generate_ci_pl()

import pandas as pd
import os
from openpyxl import load_workbook
from tqdm import tqdm
from copy import copy
from num2words import num2words

def generate_ci_pl():
    print("[INFO] Membaca file Master Order List dan Guideline Shipment...")

    # --- [FITUR] KAMUS HS CODE ---
    hs_kamus = {
        "211354": "64041990",
    }
    # -----------------------------

    try:
        # 1. Membaca Master Order List & Guideline Shipment
        df = pd.read_excel("Master Order List Adi.xlsx")
        df_db = pd.read_excel("guideline_shipment.xlsx")
    except Exception as e:
        print(f"[ERROR] Gagal membaca file excel: {e}")
        return

    # 2. Membuat folder utama Output jika belum ada
    output_dir = "Output_FI_PL_SI"
    os.makedirs(output_dir, exist_ok=True)

    # 3. GRUP PERTAMA: Mengelompokkan berdasarkan TL# (Untuk membuat Folder Gabungan)
    grouped_tl = df.groupby("TL#")
    
    print(f"[INFO] Menemukan {len(grouped_tl)} grup TL#. Memulai proses generate...\n")

    # ==============================================================================
    # JALUR 1 & 2: GENERATE COMMERCIAL INVOICE (CI) & PACKING LIST (PL)
    # ==============================================================================
    for tl_num, tl_group in tqdm(grouped_tl, desc="Proses CI & PL per TL"):
        # -- TENTUKAN NAMA FOLDER GABUNGAN --
        ci_list = sorted(tl_group["CI#"].unique().astype(str))
        all_ci_names = " & ".join(ci_list)
        tl_code_folder = tl_group.iloc[0]["TL#"]
        
        folder_name = f"{all_ci_names} - {tl_code_folder}".replace("/", "-") 
        folder_path = os.path.join(output_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        
        # SUB-LOOPING: Memproses masing-masing Invoice di dalam grup TL yang sama
        grouped_inv = tl_group.groupby("Factory Inv#")
        
        for inv_num, group in grouped_inv:
            first_row = group.iloc[0]
            date_export = first_row["EXII-FTY"]
            country = first_row["Customer Country"]
            kode_negara = first_row["Code Country"]
            tl_code = first_row["Code TL"]
            ship_by = first_row["Ship By"]
            ship_mode = first_row["MOT"]
            branch_plant = first_row["Branch Plant"]
            pod_inv_pl = first_row["POD"]
            

            # MENCARI CONSIGNEE & DELIVERY ADDRESS
            match = df_db[
                (df_db["Country"] == kode_negara) & 
                (df_db["By"] == ship_by) & 
                (df_db["Branch Plant"] == branch_plant)
            ]
            
            if not match.empty:
                consignee_data = match.iloc[0]["Consignee"]
                delivery_data = match.iloc[0]["Notify Party/Delivery Address"]
            else:
                consignee_data = "DATA TIDAK DITEMUKAN"
                delivery_data = "DATA TIDAK DITEMUKAN"

            # ------------------------------------------
            # PROSES 1: COMMERCIAL INVOICE (CI)
            # ------------------------------------------
            wb_ci = load_workbook("Template_CI_V2.xlsx")
            ws_ci = wb_ci.active
            
            ws_ci["E6"], ws_ci["G6"], ws_ci["G10"] = inv_num, date_export, country
            ws_ci["A19"], ws_ci["E19"], ws_ci["E21"] = ship_mode, ship_by, pod_inv_pl
            
            if isinstance(consignee_data, str):
                for i, baris in enumerate(consignee_data.split('\n')):
                    ws_ci[f"C{13 + i}"] = baris.strip()
            if isinstance(delivery_data, str):
                for i, baris in enumerate(delivery_data.split('\n')):
                    ws_ci[f"E{13 + i}"] = baris.strip()
            
            start_row = 26
            num_items = len(group)
            
            if num_items > 1:
                ws_ci.insert_rows(start_row + 1, amount=num_items - 1)
                for r in range(start_row + 1, start_row + num_items):
                    for c in range(1, 10): 
                        source_cell = ws_ci.cell(row=start_row, column=c)
                        target_cell = ws_ci.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font, target_cell.border, target_cell.alignment = copy(source_cell.font), copy(source_cell.border), copy(source_cell.alignment)
                            target_cell.fill, target_cell.number_format = copy(source_cell.fill), copy(source_cell.number_format)

            total_ctn, total_pairs, total_amount = 0, 0, 0
            for idx, (_, row) in enumerate(group.iterrows()):
                curr_row = start_row + idx
                ws_ci[f"A{curr_row}"] = row["PO#"]
                ws_ci[f"B{curr_row}"] = row["ProductNO"]
                ws_ci[f"C{curr_row}"] = row["EnglishName"]
                
                full_product_no = str(row["ProductNO"]).strip()
                style_only = full_product_no.split('-')[0].strip()
                ws_ci[f"D{curr_row}"] = hs_kamus.get(style_only, "64029990") 
                
                ws_ci[f"E{curr_row}"], ws_ci[f"F{curr_row}"] = row["CTN"], row["PAIRS"]
                ws_ci[f"G{curr_row}"], ws_ci[f"H{curr_row}"] = row["TE Price"], row["SCI Amount"]
                
                total_ctn += row["CTN"]
                total_pairs += row["PAIRS"]
                total_amount += row["SCI Amount"]
                
            last_row_ci = start_row + num_items + 1
            ws_ci[f"E{last_row_ci}"], ws_ci[f"F{last_row_ci}"], ws_ci[f"H{last_row_ci}"] = total_ctn, total_pairs, total_amount

            # FITUR TERBILANG CI
            terbilang_text = f"Total Amount Say US Dollar {num2words(int(total_amount), lang='en').replace('-', ' ').upper()} ONLY"
            for r in range(start_row, ws_ci.max_row + 1):
                cell_val = ws_ci[f"A{r}"].value
                if isinstance(cell_val, str):
                    teks = cell_val.strip().lower()
                    if teks == "total cartons": ws_ci[f"B{r}"] = total_ctn
                    elif teks == "total quantity": ws_ci[f"B{r}"] = total_pairs
                    elif teks == "total amount": ws_ci[f"B{r}"] = total_amount
                    elif "total amount say" in teks or "say us dollar" in teks: ws_ci[f"A{r}"] = terbilang_text

            ws_ci.print_area = f"A1:H{ws_ci.max_row - 5}"
            ws_ci.sheet_properties.pageSetUpPr.fitToPage = True
            ws_ci.page_setup.fitToWidth, ws_ci.page_setup.fitToHeight = 1, 0
            
            file_name_ci = f"{inv_num} INV {tl_code}.xlsx".replace("/", "-")
            wb_ci.save(os.path.join(folder_path, file_name_ci))
            
            # ------------------------------------------
            # PROSES 2: PACKING LIST (PL)
            # ------------------------------------------
            wb_pl = load_workbook("Template_PL_V2.xlsx")
            ws_pl = wb_pl.active
            
            ws_pl["E6"], ws_pl["G6"], ws_pl["G10"] = inv_num, date_export, country
            ws_pl["A19"], ws_pl["E19"], ws_pl["E21"] = ship_mode, ship_by, pod_inv_pl
            
            if isinstance(consignee_data, str):
                for i, baris in enumerate(consignee_data.split('\n')): ws_pl[f"C{13 + i}"] = baris.strip()
            if isinstance(delivery_data, str):
                for i, baris in enumerate(delivery_data.split('\n')): ws_pl[f"E{13 + i}"] = baris.strip()

            if num_items > 1:
                ws_pl.insert_rows(start_row + 1, amount=num_items - 1)
                for r in range(start_row + 1, start_row + num_items):
                    for c in range(1, 10): 
                        source_cell = ws_pl.cell(row=start_row, column=c)
                        target_cell = ws_pl.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font, target_cell.border, target_cell.alignment = copy(source_cell.font), copy(source_cell.border), copy(source_cell.alignment)
                            target_cell.fill, target_cell.number_format = copy(source_cell.fill), copy(source_cell.number_format)
            
            total_nw, total_gw, total_cbm = 0, 0, 0
            for idx, (_, row) in enumerate(group.iterrows()):
                curr_row = start_row + idx
                ws_pl[f"A{curr_row}"], ws_pl[f"B{curr_row}"] = row["PO#"], row["ProductNO"]
                ws_pl[f"C{curr_row}"], ws_pl[f"D{curr_row}"] = row["EnglishName"], hs_kamus.get(str(row["ProductNO"]).split('-')[0].strip(), "64029990")
                ws_pl[f"E{curr_row}"], ws_pl[f"F{curr_row}"] = row["CTN"], row["PAIRS"]
                ws_pl[f"G{curr_row}"], ws_pl[f"H{curr_row}"], ws_pl[f"I{curr_row}"] = row["N.W"], row["G.W"], row["CBM"]
                
                total_nw += row["N.W"]; total_gw += row["G.W"]; total_cbm += row["CBM"]
                
            last_row_pl = start_row + num_items + 1
            ws_pl[f"E{last_row_pl}"], ws_pl[f"F{last_row_pl}"] = total_ctn, total_pairs
            ws_pl[f"G{last_row_pl}"], ws_pl[f"H{last_row_pl}"], ws_pl[f"I{last_row_pl}"] = total_nw, total_gw, total_cbm

            for r in range(start_row, ws_pl.max_row + 1):
                cell_val = ws_pl[f"A{r}"].value
                if isinstance(cell_val, str):
                    teks = cell_val.strip().lower()
                    if teks == "total cartons": ws_pl[f"B{r}"] = total_ctn
                    elif teks == "total quantity": ws_pl[f"B{r}"] = total_pairs
                    elif teks == "total net weight": ws_pl[f"B{r}"] = total_nw
                    elif teks == "total gross weight": ws_pl[f"B{r}"] = total_gw
                    elif teks == "total cbm": ws_pl[f"B{r}"] = total_cbm

            ws_pl.print_area = f"A1:I{ws_pl.max_row - 4}"
            ws_pl.sheet_properties.pageSetUpPr.fitToPage = True
            ws_pl.page_setup.fitToWidth, ws_pl.page_setup.fitToHeight = 1, 0
            
            file_name_pl = f"{inv_num} PL {tl_code}.xlsx".replace("/", "-")
            wb_pl.save(os.path.join(folder_path, file_name_pl))

    # ==============================================================================
    # JALUR 3: GENERATE FINAL SI (FSI) - DISALIN KE MASING-MASING SUBFOLDER
    # ==============================================================================
    print("\n[INFO] CI & PL Selesai. Memulai proses generate FINAL SI (FSI)...")
    
    grouped_booking_global = df.groupby("Booking No / SO No.", sort=False)
    
    for book_no, book_group in tqdm(grouped_booking_global, desc="Proses FSI per Booking"):
        wb_si = load_workbook("Template_FSI_V2.xlsx")
        
        if "CONTAINER" in wb_si.sheetnames:
            ws_base = wb_si["CONTAINER"]
        else:
            ws_base = wb_si.active 

        start_row_si = 9
        grouped_inv_fsi = book_group.groupby("Factory Inv#", sort=False)
        
        sheets_for_inv = []
        for inv_idx in range(len(grouped_inv_fsi)):
            if inv_idx == 0:
                sheets_for_inv.append(ws_base)
            else:
                new_ws = wb_si.copy_worksheet(ws_base)
                new_ws.title = f"CONTAINER {inv_idx + 1}"
                sheets_for_inv.append(new_ws)
        
        for inv_idx, (inv_num, inv_group) in enumerate(grouped_inv_fsi):
            ws_current = sheets_for_inv[inv_idx]
            num_items_inv = len(inv_group)
            
            first_row_inv = inv_group.iloc[0]
            date_export = first_row_inv["EXII-FTY"]
            carrier = first_row_inv["FWDR"]
            contNo_sealNo = first_row_inv["Detail Cntainer no./Seal/truck no"]
            type_cont_truck = first_row_inv["Ctnr type/Truck type"]
            te_inv = first_row_inv["CI#"]
            no_pen_peb = first_row_inv["NO PEN PEB"]
            peb_date = first_row_inv["PEB DATE"]
            pod = first_row_inv["POD"]
            
            if num_items_inv > 1:
                ws_current.insert_rows(start_row_si + 1, amount=num_items_inv - 1)
                for r in range(start_row_si + 1, start_row_si + num_items_inv):
                    for c in range(1, 11): 
                        source_cell = ws_current.cell(row=start_row_si, column=c)
                        target_cell = ws_current.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.alignment = copy(source_cell.alignment)
            
            total_ctn_si, total_pairs_si, total_gw_si, total_cbm_si = 0, 0, 0, 0
            for idx, (_, row) in enumerate(inv_group.iterrows()):
                curr_row = start_row_si + idx
                ws_current[f"C{curr_row}"] = row["PO#"]
                ws_current[f"D{curr_row}"] = row["ProductNO"]
                ws_current[f"E{curr_row}"] = row["CTN"]
                ws_current[f"F{curr_row}"] = row["PAIRS"]
                ws_current[f"G{curr_row}"] = row["G.W"]
                ws_current[f"H{curr_row}"] = row["CBM"]
                ws_current[f"I{curr_row}"] = row["Branch Plant"]
                
                full_product_no = str(row["ProductNO"]).strip()
                style_only = full_product_no.split('-')[0].strip()
                ws_current[f"J{curr_row}"] = hs_kamus.get(style_only, "640299") 
                
                total_ctn_si += row["CTN"]; total_pairs_si += row["PAIRS"]
                total_gw_si += row["G.W"]; total_cbm_si += row["CBM"]
            
            last_row_si = start_row_si + num_items_inv + 6
            ws_current[f"E{last_row_si}"] = total_ctn_si     
            ws_current[f"F{last_row_si}"] = total_pairs_si   
            ws_current[f"G{last_row_si}"] = total_gw_si
            ws_current[f"H{last_row_si}"] = total_cbm_si
            
            ws_current["A1"] = f"CARRIER BOOKING# : {book_no}"
            ws_current["A4"] = inv_num
            ws_current["A6"] = carrier
            ws_current["A8"] = contNo_sealNo
            ws_current["B8"] = type_cont_truck
            ws_current["A9"] = te_inv
            ws_current["A10"] = f"EMPTY PICK-UP DATE : {pd.to_datetime(date_export).strftime('%d-%b-%y')}"
            ws_current["A11"] = f"STUFFING DATE : {pd.to_datetime(date_export).strftime('%d-%b-%y')}"
            ws_current["A12"] = f"PEB : {no_pen_peb} / {pd.to_datetime(peb_date).strftime('%d-%b-%y')}"
            ws_current["A13"] = f"POD : {pod}"
            ws_current["A14"] = f"Final Destination : {pod}"
        
        # PROSES PENYIMPANAN KE MASING-MASING SUBFOLDER TL YANG TERKAIT
        for tl_num, tl_sub_group in book_group.groupby("TL#"):
            ci_list = sorted(tl_sub_group["CI#"].unique().astype(str))
            all_ci_names = " & ".join(ci_list)
            tl_code_folder = tl_sub_group.iloc[0]["TL#"]
            
            folder_name = f"{all_ci_names} - {tl_code_folder}".replace("/", "-")
            target_folder_path = os.path.join(output_dir, folder_name)
            
            file_name_si = f"FINAL SI BC #{book_no}.xlsx".replace("/", "-")
            wb_si.save(os.path.join(target_folder_path, file_name_si))

    print("\n[SUCCESS] Semua data CI, PL, dan FSI berhasil digenerate dan didistribusikan ke subfolder!")

if __name__ == "__main__":
    generate_ci_pl()