import os
from openpyxl import load_workbook
from tqdm import tqdm
from copy import copy
from num2words import num2words
from modules.config import HS_KAMUS, OUTPUT_DIR, TPL_DIR
from openpyxl.drawing.image import Image # <--- TAMBAHKAN INI
from modules.config import HS_KAMUS, OUTPUT_DIR, TPL_DIR, BASE_DIR # Pastikan BASE_DIR ikut di-import
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D


def process_cipl(df, df_db):
    grouped_tl = df.groupby("TL#")
    print(f"[INFO] Menemukan {len(grouped_tl)} grup TL# untuk CI & PL.\n")

    for tl_num, tl_group in tqdm(grouped_tl, desc="Proses CI & PL per TL"):
        ci_list = sorted(tl_group["CI#"].unique().astype(str))
        all_ci_names = " & ".join(ci_list)
        tl_code_folder = tl_group.iloc[0]["TL#"]
        
        folder_name = f"{all_ci_names} - {tl_code_folder}".replace("/", "-") 
        folder_path = os.path.join(OUTPUT_DIR, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        
        grouped_inv = tl_group.groupby("Factory Inv#")
        
        for inv_num, group in grouped_inv:
            first_row = group.iloc[0]
            date_export = first_row["EXII-FTY"]
            country = first_row["Customer Country"]
            kode_negara = first_row["Code Country"]
            tl_code = first_row["Code TL"]
            ship_by = first_row["Ship By"]
            ship_mode = first_row["MOT"]
            branch_plant = first_row["Branch Plant"]
            pod_inv_pl = first_row["POD"]

            match = df_db[
                (df_db["Country"] == kode_negara) & 
                (df_db["By"] == ship_by) & 
                (df_db["Branch Plant"] == branch_plant)
            ]
            
            if not match.empty:
                consignee_data = match.iloc[0]["Consignee"]
                delivery_data = match.iloc[0]["Notify Party/Delivery Address"]
            else:
                consignee_data, delivery_data = "DATA TIDAK DITEMUKAN", "DATA TIDAK DITEMUKAN"

            # --- [LOGIKA TEMPLATE US] ---
            if kode_negara == "US":
                template_ci_path = os.path.join(TPL_DIR, "Template_CI_US.xlsx")
                template_pl_path = os.path.join(TPL_DIR, "Template_PL_US.xlsx")
            else:
                template_ci_path = os.path.join(TPL_DIR, "Template_CI_V2.xlsx")
                template_pl_path = os.path.join(TPL_DIR, "Template_PL_V2.xlsx")
            # ----------------------------

            # ==========================================
            # COMMERCIAL INVOICE (CI)
            # ==========================================
            wb_ci = load_workbook(template_ci_path)
            ws_ci = wb_ci.active
            
            ws_ci["E6"], ws_ci["G6"], ws_ci["G10"] = inv_num, date_export, country
            ws_ci["A19"], ws_ci["E19"], ws_ci["E21"] = ship_mode, ship_by, pod_inv_pl
            
            if isinstance(consignee_data, str):
                for i, baris in enumerate(consignee_data.split('\n')): ws_ci[f"C{13 + i}"] = baris.strip()
            if isinstance(delivery_data, str):
                for i, baris in enumerate(delivery_data.split('\n')): ws_ci[f"E{13 + i}"] = baris.strip()
            
            start_row, num_items = 26, len(group)
            
            if num_items > 1:
                ws_ci.insert_rows(start_row + 1, amount=num_items - 1)
                for r in range(start_row + 1, start_row + num_items):
                    for c in range(1, 10): 
                        source_cell, target_cell = ws_ci.cell(row=start_row, column=c), ws_ci.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font, target_cell.border, target_cell.alignment = copy(source_cell.font), copy(source_cell.border), copy(source_cell.alignment)
                            target_cell.fill, target_cell.number_format = copy(source_cell.fill), copy(source_cell.number_format)

            total_ctn, total_pairs, total_amount = 0, 0, 0
            for idx, (_, row) in enumerate(group.iterrows()):
                curr_row = start_row + idx
                ws_ci[f"A{curr_row}"], ws_ci[f"B{curr_row}"], ws_ci[f"C{curr_row}"] = row["PO#"], row["ProductNO"], row["EnglishName"]
                ws_ci[f"D{curr_row}"] = HS_KAMUS.get(str(row["ProductNO"]).split('-')[0].strip(), "64029990") 
                ws_ci[f"E{curr_row}"], ws_ci[f"F{curr_row}"] = row["CTN"], row["PAIRS"]
                ws_ci[f"G{curr_row}"], ws_ci[f"H{curr_row}"] = row["TE Price"], row["SCI Amount"]
                
                total_ctn += row["CTN"]; total_pairs += row["PAIRS"]; total_amount += row["SCI Amount"]
                
            last_row_ci = start_row + num_items + 1
            ws_ci[f"E{last_row_ci}"], ws_ci[f"F{last_row_ci}"], ws_ci[f"H{last_row_ci}"] = total_ctn, total_pairs, total_amount

            # --- [FITUR TERBILANG DOLLAR & CENTS] ---
            dollars = int(total_amount)
            cents = int(round((total_amount - dollars) * 100))
            terbilang_dollars = num2words(dollars, lang='en').replace('-', ' ').replace(',', '').upper()
            
            if cents > 0:
                terbilang_cents = num2words(cents, lang='en').replace('-', ' ').upper()
                terbilang_text = f"Total Amount Say US Dollar {terbilang_dollars} AND CENTS {terbilang_cents} ONLY"
            else:
                terbilang_text = f"Total Amount Say US Dollar {terbilang_dollars} ONLY"
            for r in range(start_row, ws_ci.max_row + 1):
                if isinstance(ws_ci[f"A{r}"].value, str):
                    teks = ws_ci[f"A{r}"].value.strip().lower()
                    if teks == "total cartons": ws_ci[f"B{r}"] = total_ctn
                    elif teks == "total quantity": ws_ci[f"B{r}"] = total_pairs
                    elif teks == "total amount": ws_ci[f"B{r}"] = total_amount
                    elif "total amount say" in teks or "say us dollar" in teks: ws_ci[f"A{r}"] = terbilang_text
            
            # ==========================================
            # --- [FITUR STAMP - DINAMIS & SUB-CELL PIXEL OFFSET] ---
            stamp_path = os.path.join(BASE_DIR, "stamp.png")
            
            if os.path.exists(stamp_path):
                img_stamp = Image(stamp_path)
                
                # 1. PENGATURAN UKURAN STAMP
                img_stamp.width = 360  
                img_stamp.height = 100 
                
                target_cell_found = False
                
                # 2. MENCARI ANCHOR "Dennis Lin"
                for r in range(start_row, ws_ci.max_row + 1):
                    for c in range(1, 10): 
                        cell = ws_ci.cell(row=r, column=c)
                        val = cell.value
                        
                        if isinstance(val, str) and "dennis lin" in val.strip().lower():
                            # --- 3. PENGATURAN SEL JANGKAR ---
                            row_offset = -6 
                            col_offset = -1   
                            
                            # Menghitung index sel jangkar (Openpyxl AnchorMarker mulai dari index 0)
                            anchor_col = (cell.column - 1) + col_offset
                            anchor_row = (cell.row - 1) + row_offset
                            
                            # --- 4. PENGATURAN GESER PIXEL BEBAS (KUNCI PRESISI) ---
                            # Silakan ubah angka ini sampai jaraknya pas!
                            geser_x_px = -35  # (+) untuk geser kanan, (-) untuk geser kiri
                            geser_y_px = 15  # (+) untuk turun ke bawah, (-) untuk naik ke atas
                            
                            # Membuat koordinat custom EMU
                            marker = AnchorMarker(
                                col=anchor_col, 
                                colOff=pixels_to_EMU(geser_x_px), 
                                row=anchor_row, 
                                rowOff=pixels_to_EMU(geser_y_px)
                            )
                            size = XDRPositiveSize2D(
                                cx=pixels_to_EMU(img_stamp.width), 
                                cy=pixels_to_EMU(img_stamp.height)
                            )
                            
                            # Menerapkan koordinat tersebut ke gambar
                            img_stamp.anchor = OneCellAnchor(_from=marker, ext=size)
                            target_cell_found = True
                            break
                    
                    if target_cell_found:
                        break
                        
                # 5. TEMPEL GAMBAR
                if target_cell_found:
                    # Kita TIDAK butuh kordinat seperti "G40" lagi karena sudah diatur oleh img_stamp.anchor
                    ws_ci.add_image(img_stamp) 
            # ==========================================

            ws_ci.print_area = f"A1:H{ws_ci.max_row - 5}"
            ws_ci.sheet_properties.pageSetUpPr.fitToPage = True
            ws_ci.page_setup.fitToWidth, ws_ci.page_setup.fitToHeight = 1, 0
            wb_ci.save(os.path.join(folder_path, f"{inv_num} INV {tl_code}.xlsx".replace("/", "-")))
            
            # ==========================================
            # PACKING LIST (PL)
            # ==========================================
            wb_pl = load_workbook(template_pl_path)
            ws_pl = wb_pl.active
            
            ws_pl["E6"], ws_pl["G6"], ws_pl["G10"] = inv_num, date_export, country
            ws_pl["A19"], ws_pl["E19"], ws_pl["E21"] = ship_mode, ship_by, pod_inv_pl
            
            if isinstance(consignee_data, str):
                for i, baris in enumerate(consignee_data.split('\n')): ws_pl[f"C{13 + i}"] = baris.strip()
            if isinstance(delivery_data, str):
                for i, baris in enumerate(delivery_data.split('\n')): ws_pl[f"E{13 + i}"] = baris.strip()

            if num_items > 1:
                ws_pl.insert_rows(start_row + 1, amount=num_items - 1)
                for r in range(start_row + 1, start_row + num_items):
                    for c in range(1, 10): 
                        source_cell, target_cell = ws_pl.cell(row=start_row, column=c), ws_pl.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font, target_cell.border, target_cell.alignment = copy(source_cell.font), copy(source_cell.border), copy(source_cell.alignment)
                            target_cell.fill, target_cell.number_format = copy(source_cell.fill), copy(source_cell.number_format)
            
            total_nw, total_gw, total_cbm = 0, 0, 0
            for idx, (_, row) in enumerate(group.iterrows()):
                curr_row = start_row + idx
                ws_pl[f"A{curr_row}"], ws_pl[f"B{curr_row}"], ws_pl[f"C{curr_row}"] = row["PO#"], row["ProductNO"], row["EnglishName"]
                ws_pl[f"D{curr_row}"] = HS_KAMUS.get(str(row["ProductNO"]).split('-')[0].strip(), "64029990")
                ws_pl[f"E{curr_row}"], ws_pl[f"F{curr_row}"] = row["CTN"], row["PAIRS"]
                ws_pl[f"G{curr_row}"], ws_pl[f"H{curr_row}"], ws_pl[f"I{curr_row}"] = row["N.W"], row["G.W"], row["CBM"]
                
                total_nw += row["N.W"]; total_gw += row["G.W"]; total_cbm += row["CBM"]
                
            last_row_pl = start_row + num_items + 1
            ws_pl[f"E{last_row_pl}"], ws_pl[f"F{last_row_pl}"] = total_ctn, total_pairs
            ws_pl[f"G{last_row_pl}"], ws_pl[f"H{last_row_pl}"], ws_pl[f"I{last_row_pl}"] = total_nw, total_gw, total_cbm

            for r in range(start_row, ws_pl.max_row + 1):
                if isinstance(ws_pl[f"A{r}"].value, str):
                    teks = ws_pl[f"A{r}"].value.strip().lower()
                    if teks == "total cartons": ws_pl[f"B{r}"] = total_ctn
                    elif teks == "total quantity": ws_pl[f"B{r}"] = total_pairs
                    elif teks == "total net weight": ws_pl[f"B{r}"] = total_nw
                    elif teks == "total gross weight": ws_pl[f"B{r}"] = total_gw
                    elif teks == "total cbm": ws_pl[f"B{r}"] = total_cbm
                    
           # ==========================================
            # --- [FITUR STAMP - DINAMIS & SUB-CELL PIXEL OFFSET] ---
            stamp_path = os.path.join(BASE_DIR, "stamp.png")
            
            if os.path.exists(stamp_path):
                img_stamp = Image(stamp_path)
                
                # 1. PENGATURAN UKURAN STAMP
                img_stamp.width = 360  
                img_stamp.height = 100 
                
                target_cell_found = False
                
                # 2. MENCARI ANCHOR "Dennis Lin"
                for r in range(start_row, ws_pl.max_row + 1):
                    for c in range(1, 10): 
                        cell = ws_pl.cell(row=r, column=c)
                        val = cell.value
                        
                        if isinstance(val, str) and "dennis lin" in val.strip().lower():
                            # --- 3. PENGATURAN SEL JANGKAR ---
                            row_offset = -6 
                            col_offset = -1   
                            
                            # Menghitung index sel jangkar (Openpyxl AnchorMarker mulai dari index 0)
                            anchor_col = (cell.column - 1) + col_offset
                            anchor_row = (cell.row - 1) + row_offset
                            
                            # --- 4. PENGATURAN GESER PIXEL BEBAS (KUNCI PRESISI) ---
                            # Silakan ubah angka ini sampai jaraknya pas!
                            geser_x_px = -25  # (+) untuk geser kanan, (-) untuk geser kiri
                            geser_y_px = 15  # (+) untuk turun ke bawah, (-) untuk naik ke atas
                            
                            # Membuat koordinat custom EMU
                            marker = AnchorMarker(
                                col=anchor_col, 
                                colOff=pixels_to_EMU(geser_x_px), 
                                row=anchor_row, 
                                rowOff=pixels_to_EMU(geser_y_px)
                            )
                            size = XDRPositiveSize2D(
                                cx=pixels_to_EMU(img_stamp.width), 
                                cy=pixels_to_EMU(img_stamp.height)
                            )
                            
                            # Menerapkan koordinat tersebut ke gambar
                            img_stamp.anchor = OneCellAnchor(_from=marker, ext=size)
                            target_cell_found = True
                            break
                    
                    if target_cell_found:
                        break
                        
                # 5. TEMPEL GAMBAR
                if target_cell_found:
                    # Kita TIDAK butuh kordinat seperti "G40" lagi karena sudah diatur oleh img_stamp.anchor
                    ws_pl.add_image(img_stamp) 
            # ==========================================
            
            ws_pl.print_area = f"A1:I{ws_pl.max_row - 4}"
            ws_pl.sheet_properties.pageSetUpPr.fitToPage = True
            ws_pl.page_setup.fitToWidth, ws_pl.page_setup.fitToHeight = 1, 0
            wb_pl.save(os.path.join(folder_path, f"{inv_num} PL {tl_code}.xlsx".replace("/", "-")))