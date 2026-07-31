import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from copy import copy
from modules.config import HS_KAMUS, OUTPUT_DIR, TPL_DIR

def process_fsi(df):
    grouped_booking_global = df.groupby("Booking No / SO No.", sort=False)
    template_fsi_path = os.path.join(TPL_DIR, "Template_FSI_V2.xlsx")
    
    for book_no, book_group in tqdm(grouped_booking_global, desc="Proses FSI per Booking"):
        wb_si = load_workbook(template_fsi_path)
        ws_base = wb_si["CONTAINER"] if "CONTAINER" in wb_si.sheetnames else wb_si.active 

        start_row_si = 9
        grouped_inv_fsi = book_group.groupby("Factory Inv#", sort=False)
        
        sheets_for_inv = []
        for inv_idx in range(len(grouped_inv_fsi)):
            if inv_idx == 0:
                sheets_for_inv.append(ws_base)
            else:
                new_ws = wb_si.copy_worksheet(ws_base)
                new_ws.title = f"CONTAINER {inv_idx + 1}"
                sheets_for_inv.append(new_ws)
        
        for inv_idx, (inv_num, inv_group) in enumerate(grouped_inv_fsi):
            ws_current = sheets_for_inv[inv_idx]
            num_items_inv = len(inv_group)
            
            first_row_inv = inv_group.iloc[0]
            date_export = first_row_inv["EXII-FTY"]
            carrier = first_row_inv["FWDR"]
            contNo_sealNo = first_row_inv["Detail Cntainer no./Seal/truck no"]
            type_cont_truck = first_row_inv["Ctnr type/Truck type"]
            te_inv = first_row_inv["CI#"]
            no_pen_peb = first_row_inv["NO PEN PEB"]
            peb_date = first_row_inv["PEB DATE"]
            pod = first_row_inv["POD"]
            
            if num_items_inv > 1:
                ws_current.insert_rows(start_row_si + 1, amount=num_items_inv - 1)
                for r in range(start_row_si + 1, start_row_si + num_items_inv):
                    for c in range(3, 11): 
                        source_cell, target_cell = ws_current.cell(row=start_row_si, column=c), ws_current.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font, target_cell.border, target_cell.alignment = copy(source_cell.font), copy(source_cell.border), copy(source_cell.alignment)
                            target_cell.fill, target_cell.number_format = copy(source_cell.fill), copy(source_cell.number_format)
            
            total_ctn_si, total_pairs_si, total_gw_si, total_cbm_si = 0, 0, 0, 0
            for idx, (_, row) in enumerate(inv_group.iterrows()):
                curr_row = start_row_si + idx
                ws_current[f"C{curr_row}"], ws_current[f"D{curr_row}"] = row["PO#"], row["ProductNO"]
                ws_current[f"E{curr_row}"], ws_current[f"F{curr_row}"] = row["CTN"], row["PAIRS"]
                ws_current[f"G{curr_row}"], ws_current[f"H{curr_row}"] = row["G.W"], row["CBM"]
                ws_current[f"I{curr_row}"] = row["Branch Plant"]
                
                ws_current[f"J{curr_row}"] = HS_KAMUS.get(str(row["ProductNO"]).split('-')[0].strip(), "640299") 
                
                total_ctn_si += row["CTN"]; total_pairs_si += row["PAIRS"]
                total_gw_si += row["G.W"]; total_cbm_si += row["CBM"]
            
            last_row_si = start_row_si + num_items_inv + 6
            ws_current[f"E{last_row_si}"], ws_current[f"F{last_row_si}"] = total_ctn_si, total_pairs_si
            ws_current[f"G{last_row_si}"], ws_current[f"H{last_row_si}"] = total_gw_si, total_cbm_si
            
            ws_current["A1"] = f"CARRIER BOOKING# : {book_no}"
            ws_current["A4"], ws_current["A6"] = inv_num, carrier
            ws_current["A8"], ws_current["B8"] = contNo_sealNo, type_cont_truck
            ws_current["A9"] = te_inv
            ws_current["A10"] = f"EMPTY PICK-UP DATE : {pd.to_datetime(date_export).strftime('%d-%b-%y')}"
            ws_current["A11"] = f"STUFFING DATE : {pd.to_datetime(date_export).strftime('%d-%b-%y')}"
            ws_current["A12"] = f"PEB : {no_pen_peb} / {pd.to_datetime(peb_date).strftime('%d-%b-%y')}"
            ws_current["A13"] = ws_current["A14"] = f"POD : {pod}" if "A13" else f"Final Destination : {pod}" # A14 fix
            ws_current["A14"] = f"Final Destination : {pod}"
        
        for tl_num, tl_sub_group in book_group.groupby("TL#"):
            ci_list = sorted(tl_sub_group["CI#"].unique().astype(str))
            all_ci_names = " & ".join(ci_list)
            tl_code_folder = tl_sub_group.iloc[0]["TL#"]
            
            folder_name = f"{all_ci_names} - {tl_code_folder}".replace("/", "-")
            target_folder_path = os.path.join(OUTPUT_DIR, folder_name)
            os.makedirs(target_folder_path, exist_ok=True)
            
            file_name_si = f"FINAL SI BC #{book_no}.xlsx".replace("/", "-")
            wb_si.save(os.path.join(target_folder_path, file_name_si))